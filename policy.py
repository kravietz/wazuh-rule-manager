# -*- coding: utf-8 -*-
import json
import pathlib
import re
from collections import OrderedDict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

from colors import C

__author__ = 'Pawel Krawczyk'


class PolicyEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, pathlib.Path):
            return obj.name
        return obj.__dict__


class Policy:
    """
    Class that represents an abstract Wazuh policy incorporating all the rules. Rules are organized in Collections
    and reflect the customary organization of Wazuh rules into separate XML files.
    """

    def __init__(self) -> None:
        self.filename = None
        self.rules = OrderedDict()

    def from_file(self, filename: Path) -> None:
        self.filename = filename
        self._load()

    def from_rules(self, rules_manager: object) -> None:
        """
        Create Policy object from XML parsed rules passed as RuleManager object
        """
        # top-level import causes circular reference
        from manager import RuleManager

        # just a sanity check
        if not isinstance(rules_manager, RuleManager):
            raise ValueError('Need RuleManager instance here')

        for input_collection in rules_manager.collections:
            collection = self.Collection(input_collection.filename)
            for input_rule in input_collection.get_all_rules():
                # input_rule is Element('<rule>')
                rule_contents = dict()
                for field in input_rule.getchildren():
                    # field is consecutive inner field, such as Element('<description>')
                    # note there is loss of information here - we are discarding attributes completely
                    # however in the Policy object we only really care about `level` and `id`, all the
                    # other fields are purely informative to assist rule review and adjustment
                    rule_contents[field.tag] = field.text

                # populate fields that do not come as Element('<rule>') subfields
                rule_contents['collection'] = collection  # created above
                rule_contents['level'] = input_rule.get('level')  # rule attribute
                rule_contents['id'] = input_rule.get('id')  # rule attribute

                # initialize a new Rule object from the XML fields
                new_rule = self.Rule(rule_contents)
                self.rules[new_rule.id] = new_rule

    def get_rules_by_collection(self, collection: object) -> list:
        """
        Return all rules in specified Collection sorted by id (ascending)
        """
        ret = []
        for rule in self.rules.values():
            if rule.collection == collection:
                ret.append(rule)
        return sorted(ret, key=lambda k: k.id)

    def write(self, output_file: Path) -> None:
        """
        Writes the policy into an XLSX spreadsheet
        """
        workbook = openpyxl.Workbook(write_only=True)
        for collection in self.get_collections(sort=True):
            worksheet = workbook.create_sheet(title=collection.filename, index=collection.priority)
            header_row = True
            header_fields = []  # for code analysis
            for rule in self.get_rules_by_collection(collection):

                if header_row:
                    header_fields = rule.as_header()
                    worksheet.append(header_fields)
                    header_row = False

                # do some purely visual field tossing - put level and id into the front
                row = []
                for field_name in header_fields:
                    # this is not a rule field, this is internal data
                    if field_name == 'collection':
                        continue
                    # manually create duplicate of `level` for visual comparison in the spreadsheet
                    elif field_name == 'prev_level':
                        row.append(rule.__dict__['level'])
                    else:
                        # some cells are empty in given column, this is not a problem
                        row.append(rule.__dict__.get(field_name))

                # fill in the complete row into the worksheet
                worksheet.append(row)

            # highlight the header row
            worksheet.row_dimensions[1].font = Font(bold=True)

        # finally save the whole workbook file
        workbook.save(str(output_file))

    class Collection:
        """
        Rule collection reflecting one rules file such as `0016-wazuh_rules.xml`. Out of that we derive
        name (`wazuh`) and priority (`16`) which are later used in writing the rules back to disk.
        """

        def __init__(self, filename: Path):
            self.filename = str(filename)

            # generate collection name out of the filename
            if re.match(r'^\d+-[\w-]+\.xml$', filename):
                self.name = filename.split('-')[1].replace('.xml', '').replace('_rules', '')
                self.priority = int(filename.split('-')[0])
            elif re.match(r'^[\w-]+\.xml$', filename):
                self.name = filename.replace('.xml', '').replace('_rules', '')
            else:
                raise ValueError('Collection name does not follow the 0000-name_rules.xml convention', filename)

        def __str__(self):
            return self.filename

    class Rule:
        """
        A single Wazuh rule object.
        """

        def __init__(self, kv: dict):
            # only to stop code inspection alerts about non-existent field references
            self.id = None
            self.level = 0

            core_fields = {'id', 'collection', 'level'}
            if not core_fields.issubset(kv.keys()):
                raise ValueError('ERROR: rule without mandatory fields', core_fields, ', rule=', kv)

            for k, v in kv.items():
                if k is None:
                    print(C.Y, 'WARNING:', C.X, 'Ignoring cell in collection', kv['collection'].filename,
                          'without a heading value=', v)
                    print(kv)
                    continue
                self.__dict__[k] = v

        def as_header(self):
            fields = self.__dict__.copy()
            ret = ['id', 'prev_level', 'level']
            del fields['id']
            del fields['level']
            ret += fields.keys()
            return ret

        def __str__(self):
            return str(self.__dict__)

    def get_collections(self, sort: bool = False) -> list:
        """
        Return a list of all Collections sorted by their priority (ascending)
        """
        ret = set()
        for rule in self.rules.values():
            ret.add(rule.collection)
        if sort:
            return sorted(ret, key=lambda k: k.priority)
        else:
            return ret

    def fixup(self) -> None:
        priorities = []
        for collection in self.get_collections():
            if hasattr(collection, 'priority'):
                priorities.append(collection.priority)
        last_priority = max(priorities)
        for collection in self.get_collections():
            if not hasattr(collection, 'priority'):
                last_priority += 100
                collection.priority = last_priority
                new_filename = '{:04d}-{}_rules.xml'.format(collection.priority, collection.name)
                print(C.Y, 'WARNING:', C.X, 'Collection', C.H, collection, C.X,
                      'had no priority, assigning first available',
                      last_priority, 'and renaming to', C.H, new_filename, C.X)
                collection.filename = new_filename

    def num_collections(self) -> int:
        return len(self.get_collections())

    def num_rules(self) -> int:
        return len(self.rules)

    def get_rule_by_id(self, rule_id: int) -> Rule:
        try:
            return self.rules.get(int(rule_id))
        except (KeyError, ValueError):
            try:
                return self.rules[rule_id]
            except KeyError:
                return self.rules.get(str(rule_id))

    def _load(self) -> None:

        policy_workbook = openpyxl.load_workbook(str(self.filename), read_only=True)

        for collection_filename in policy_workbook.sheetnames:

            if not collection_filename.endswith('.xml'):
                print(C.Y, 'WARNING:', C.X, 'Skipping collection', collection_filename,
                      'because its name doesn\'t end with .xml')
                continue

            rule_worksheet = policy_workbook[collection_filename]

            collection = self.Collection(collection_filename)

            # names of header fields for this collection used to map fields into rule
            rule_collection_headers = []
            in_heading_row = True  # type: bool

            for row in rule_worksheet.rows:

                rule_contents = {}

                col_idx = 0

                for cell in row:

                    if in_heading_row:
                        rule_collection_headers.append(cell.value)
                        continue

                    if cell.value is not None:
                        field_name = rule_collection_headers[col_idx]
                        rule_contents[field_name] = cell.value

                    col_idx += 1

                if in_heading_row:
                    in_heading_row = False
                    continue

                # reading cells completed, produce Rule object
                if not {'id', 'level'}.issubset(rule_contents.keys()):
                    print(C.Y, 'WARNING:', C.X, 'collection', collection.filename, ' has row without id and level',
                          rule_contents)
                    print(row)
                    continue

                # append collection information
                rule_contents["collection"] = collection

                new_rule = self.Rule(rule_contents)
                if new_rule.id in self.rules:
                    raise IndexError('Duplicate rule id={}, new rule={}\n old rule={}'.format(new_rule.id, new_rule,
                                                                                              self.rules[new_rule.id]))
                self.rules[new_rule.id] = new_rule
